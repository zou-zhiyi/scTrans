import datetime
import glob
import random
import seaborn as sns
import pickle
import os
from collections import Counter

import sklearn.metrics
from sklearn.cluster import KMeans
from typing import List

import numpy as np
import torch
import scanpy as sc
import anndata as ad
import umap
import matplotlib.pyplot as plt
import xlrd
import xlwt
from matplotlib_venn import venn2, venn3
from openpyxl.reader.excel import load_workbook
from sklearn.decomposition import PCA
from sklearn.metrics import adjusted_rand_score, accuracy_score, f1_score, silhouette_score, precision_score
from sklearn.model_selection import train_test_split
from sklearn import manifold
from torch import nn
import copy
import pandas as pd
import gc
import ast


class WordIdxDic:

    def __init__(self):
        self.word2idx_dic = {}
        self.idx2word_dic = {}
        self.current_idx = 0

    def insert(self, gene):
        if gene in self.word2idx_dic.keys():
            return
        else:
            while self.current_idx in self.idx2word_dic.keys():
                self.current_idx += 1
            self.word2idx_dic[gene] = self.current_idx
            self.idx2word_dic[self.current_idx] = gene

    def getGene(self, idx):
        return self.idx2word_dic.get(idx, None)

    def getIdx(self, gene):
        return self.word2idx_dic.get(gene, None)


def write_file_to_pickle(data, save_path):
    with open(save_path, 'wb') as file_to_write:
        pickle.dump(data, file_to_write)


def read_file_from_pickle(save_path):
    with open(save_path, 'rb') as file_to_read:
        data = pickle.load(file_to_read)
    return data


def merger_gene_dic(adata: sc.AnnData, gene_idx_dic=None) -> WordIdxDic:
    if gene_idx_dic is None:
        gene_idx_dic = WordIdxDic()
    for gene in adata.var_names:
        gene_idx_dic.insert(gene.lower())
    if 'cell_type' in adata.obs.keys():
        for cell_type in set(adata.obs['cell_type']):
            gene_idx_dic.insert(str(cell_type).lower())
    if adata.uns.get('tissues', None) is not None:
        for tissue in adata.uns['tissues']:
            gene_idx_dic.insert(tissue.lower())
    return gene_idx_dic


def merger_gene_dic_from_varname(var_names, cell_types, gene_idx_dic=None) -> WordIdxDic:
    if gene_idx_dic is None:
        gene_idx_dic = WordIdxDic()
    for gene in var_names:
        if isinstance(gene, str):
            gene_idx_dic.insert(gene.lower())
        else:
            gene_idx_dic.insert(gene)
    for cell_type in set(cell_types):
        if isinstance(cell_type, str):
            gene_idx_dic.insert(cell_type.lower())
        else:
            gene_idx_dic.insert(cell_type)
    # if adata.uns.get('tissues', None) is not None:
    #     for tissue in adata.uns['tissues']:
    #         gene_idx_dic.insert(tissue.lower())
    return gene_idx_dic

def read_data_from_csv(data_path, cell_type_path, dataset_prefix, check_data=False):
    rna_data = pd.read_csv(data_path, header=None, low_memory=False)
    cell_type_data = pd.read_csv(cell_type_path, header=None, low_memory=False)
    adata = sc.AnnData(np.array(rna_data.iloc[1:, 1:].transpose(), dtype=np.float32))
    # print(adata.shape)
    # print(rna_data.iloc[1:, 0])
    # print(rna_data.iloc[1:, 0].shape)
    adata.var['gene_name'] = np.array(rna_data.iloc[1:, 0])
    adata.var_names = np.array(rna_data.iloc[1:, 0])
    adata.obs_names = np.array(rna_data.iloc[0, 1:])
    print(dataset_prefix)
    adata.obs['cell_name'] = np.array(rna_data.iloc[0, 1:] + dataset_prefix)
    # print(cell_type_data.iloc[:, 2])
    adata.obs['cell_type'] = np.array(cell_type_data.iloc[1:, 2])

    if check_data:
        check_anndata_direct(adata)
    return adata
    # cell_type_data = pd.read_csv(cell_type_path, header=None, low_memory=False)
    # check_anndata_direct(adata)
    # write_to_h5ad(adata, f'Bone_h5/{data_path}')



def label_transform(adata: sc.AnnData, filepath):
    cell_type_set = set(adata.obs['cell_type'])
    cell_type_dic = {}
    cnt = 0
    for ct in cell_type_set:
        cell_type_dic[ct] = cnt
        cnt += 1
    adata.obs['cell_type_idx'] = adata.obs['cell_type'].map(lambda x: cell_type_dic[x])
    # adata.uns['cell_type_nums'] = cnt
    # adata.uns['cell_type_dic'] = cell_type_dic
    check_anndata_direct(adata)
    print(f"start to write: {filepath}")
    print(f'before gene nums:{len(adata.var_names)}')
    sc.pp.filter_genes(adata, min_cells=1)
    print(f'after gene nums:{len(adata.var_names)}')
    write_to_h5ad(adata, filepath)


def merge_files(file_prefixes, save_filename, tissue_name):
    cell_nums = 0
    gene_set = set()
    cell_type_set = set()
    total_adata = None
    if len(file_prefixes) == 0:
        return
    # print(data_files)
    for file_prefix in file_prefixes:
        print(file_prefix)
        data_files = glob.glob(f'{file_prefix}*_data.csv')
        cnt = 0
        for data_file in data_files:
            father_path = os.path.abspath((os.path.dirname(data_file)))
            # print(father_path)
            # print(data_file)

            cell_type_file = data_file.split(os.sep)[-1].split('_')
            cell_type_file[-1] = 'celltype.csv'
            cell_type_file = '_'.join(cell_type_file)
            print(f'{cell_type_file}')
            adata = read_data_from_csv(data_file, f'{father_path}{os.sep}{cell_type_file}', '_' + cell_type_file, True)
            cell_nums += len(adata.obs)
            gene_set.update(set(np.array(adata.var['gene_name'])))
            cell_type_set.update(set(np.array(adata.obs['cell_type'])))
            adata.obs['tissues'] = tissue_name
            adata.obs['batch_id'] = cnt
            cnt += 1
            if total_adata is None:
                total_adata = adata
            else:
                total_adata = total_adata.concatenate(adata, join='outer', fill_value=0, uns_merge="first")
            gc.collect()
        print(f'cell nums: {cell_nums}')
        print(f'gene set nums: {len(gene_set)}')
        print(f'cell type set nums: {len(cell_type_set)}')
        print(cell_type_set)
        if total_adata.uns.get('tissues', None) is None:
            total_adata.uns['tissues'] = [tissue_name]
        else:
            total_adata.uns['tissues'].append(tissue_name)
    sc.pp.highly_variable_genes(total_adata)
    label_transform(total_adata, save_filename)
    check_anndata_direct(total_adata)
    print(total_adata.obs["cell_type_idx"])
    print(set(total_adata.obs["cell_type_idx"]))
    return total_adata

def merge_multi_files(fileprefx, save_filename, tissue_name=None):
    if isinstance(fileprefx, list):
        return merge_files(fileprefx, save_filename, tissue_name)
    else:
        return merge_files([fileprefx], save_filename, tissue_name)

def loss_visual(loss_total, test_loss_total, idx=''):
    plt.cla()
    plt.clf()
    y = loss_total
    print(loss_total)
    if not os.path.exists('loss'):
        os.mkdir('loss')
    x = [i for i in range(len(y))]
    plt.plot(x, y)
    x = [i for i in range(len(test_loss_total))]
    plt.plot(x, test_loss_total)
    plt.xlabel("epoch")
    plt.ylabel("loss")
    plt.grid()
    if not os.path.exists('loss'):
        os.mkdir('loss')
    plt.savefig('loss/loss_' + str(idx) + '_' + datetime.datetime.today().strftime("%Y_%m_%d_%H_%M_%S") + '.jpg')
    # plt.show()
    print('plt saved')
    plt.close()


def randomcolor():
    colorArr = ['1', '2', '3', '4', '5', '6', '7', '8', '9', 'A', 'B', 'C', 'D', 'E', 'F']
    color = "#" + ''.join([random.choice(colorArr) for i in range(6)])
    return color


def randommarker():
    marker_list = ['o', 'v', '^', '<', '>', 's', 'p', '*', 'D', 'P', 'X']
    marker = random.choice(marker_list)
    return marker


def umap_plot(data, label_name, save_file_name):
    plt.figure(figsize=(10, 10), dpi=300)
    reducer = umap.UMAP()
    embedding = reducer.fit_transform(data)
    # nan_error_value = adata.obs['cell_type_idx'].min()
    # max_type_value = adata.obs['cell_type_idx'].max()
    # adata.obs['cell_type_idx'][adata.obs['cell_type_idx'] == nan_error_value] = 0
    # target = adata.obs['cell_type_idx']
    # plt.subplots(300)
    # print(embbeding.shape)
    label_set = set(label_name.tolist())
    cnt = 0
    for l in label_set:
        tmp = (label_name == l)
        # print(tmp.shape)
        # print(tmp.sum())
        plt.scatter(embedding[tmp, 0], embedding[tmp, 1], marker='o', c=randomcolor(), s=5, label=l)
        cnt += 1
    plt.legend(loc="upper right", title="Classes")

    # legend = ax.legend(*scatter.legend_elements(),loc="lower left", title="Classes")
    # ax.add_artist(legend)
    plt.savefig(save_file_name)
    plt.show()
    plt.close()


def clones(module, N):
    "Produce N identical layers."
    return nn.ModuleList([copy.deepcopy(module) for _ in range(N)])


def tsne_plot(data, label_name, save_file_name, color_map=None, title_name=''):
    if color_map is None:
        color_map = {}

    reducer = manifold.TSNE(n_components=2, init='pca', random_state=1)
    embedding = reducer.fit_transform(data)
    # nan_error_value = adata.obs['cell_type_idx'].min()
    # max_type_value = adata.obs['cell_type_idx'].max()
    # adata.obs['cell_type_idx'][adata.obs['cell_type_idx'] == nan_error_value] = 0
    # target = adata.obs['cell_type_idx']
    # plt.subplots(300)
    # print(embedding)
    # print(embedding.shape)
    if not isinstance(label_name, list):
        label_name = label_name.tolist()
    label_name = np.array(label_name)
    label_set = set(label_name)
    return scatter_plot(label_set, label_name, color_map, embedding, title_name, save_file_name)





def scatter_plot(label_set, label_name, color_map, embedding, title_name, save_file_name):
    if color_map is None:
        color_map = {}
    plt.figure(figsize=(10, 10), dpi=300)
    print(label_set)
    cnt = 0
    for l in label_set:
        tmp = (label_name == l)
        # print(tmp)
        # print(tmp.shape)
        # print(tmp.sum())
        if l in color_map.keys():
            color, marker = color_map[l]
        else:
            color, marker = randomcolor(), randommarker()
            color_map[l] = (color, marker)
        plt.scatter(embedding[tmp, 0], embedding[tmp, 1], marker=marker, c=color, s=5, label=l)

        cnt += 1
        # print(cnt)
    plt.title(title_name, fontsize=80)
    plt.legend(loc="upper right", bbox_to_anchor=(1.25, 1))
    # legend = ax.legend(*scatter.legend_elements(),loc="lower left", title="Classes")
    # ax.add_artist(legend)
    plt.savefig(save_file_name, overwrite=True, bbox_inches='tight')
    plt.show()
    plt.close()
    return color_map


def leiden_clustering(embedding, n_clusters=10):
    # adata = sc.AnnData(X=embedding)
    # sc.pp.neighbors(adata)
    # sc.pp.k
    # sc.tl.leiden(adata)

    esimator = KMeans(n_clusters=n_clusters)
    esimator.fit(embedding)
    labels = esimator.labels_
    # sc.tl.louvain(adata)
    # sc.tl.tsne(adata, random_state=1, use_fast_tsne=False)
    # print(adata.obs['leiden'])
    # return adata.obs['louvain'].tolist()
    # labels = adata.obs['leiden'].tolist()
    si = silhouette_score(X=embedding, labels=labels)
    return labels, si
    # sc.pl.tsne(adata, color=['leiden'], save=save_path)

    # esimator = KMeans(n_clusters=10)
    # esimator.fit(embedding)
    # return esimator.labels_


def tsne_plot_heat(data, label_list, label_name, save_file_name,  title_name='', cmap='viridis', figsize=(8 / 2.54, 8 / 2.54)):
    my_dpi = 300
    plt.figure(figsize=figsize, dpi=my_dpi)
    plt.rcParams['font.family'] = 'Arial'
    plt.rcParams.update({"font.size": 6})
    # plt.figure(dpi=300)
    reducer = manifold.TSNE(n_components=2, init='pca', random_state=1)
    embedding = reducer.fit_transform(data)

    plt.xticks([])
    plt.yticks([])
    plt.scatter(embedding[:, 0], embedding[:, 1], marker='o', c=label_list, s=0.25, cmap=cmap, label=label_name,
                rasterized=True)
    # plt.scatter(embedding[:, 0], embedding[:, 1], marker='o', c=label_list, s=0.25, cmap='OrRd', label=label_name)

    plt.colorbar()

    plt.title(title_name, fontsize=6)
    plt.savefig(save_file_name)
    plt.show()
    plt.close()


def check_anndata_direct(data):
    print("Data matrix:")
    print(data.shape)
    print(data.X)
    print("=======================")
    print("Data obs:")
    print(data.obs)
    print("=======================")
    print("Data obs keys")
    print(data.obs.keys())
    print("=======================")
    print("Data var:")
    print(data.var)
    print("=======================")
    print("Data var keys")
    print(data.var.keys())
    print("=======================")
    print("Data uns:")
    print(data.uns)
    print("=======================")


def check_anndata(filepath, is_print=False):
    data = ad.read_h5ad(filepath)
    if is_print:
        print("Data matrix:")
        print(data.shape)
        print(data.X)
        print("=======================")
        print("Data obs:")
        print(data.obs)
        print("=======================")
        print("Data obs keys")
        print(data.obs.keys())
        print("=======================")
        print("Data var:")
        print(data.var)
        print("=======================")
        print("Data var keys")
        print(data.var.keys())
        print("=======================")
        print("Data uns:")
        print(data.uns)
        print("=======================")
    return data


def write_to_h5ad(anndata, filepath, copy=False):
    anndata.write_h5ad(filepath)
    if copy:
        anndata.write_h5ad(filepath + "_copy")


def stratify_split(total_data, random_seed=None, test_size=0.1, label_list=None):
    # print(Counter(total_data.obs['Cluster']))
    if label_list is not None:
        train_data, test_data = train_test_split(total_data, stratify=label_list, test_size=test_size,
                                                 random_state=random_seed)
    else:
        train_data, test_data = train_test_split(total_data, test_size=test_size, random_state=random_seed)
    # print(Counter(test_data.obs['Cluster']))
    return train_data, test_data


def calculate_score(true_label, pred_label):
    ari = adjusted_rand_score(true_label, pred_label)
    acc = accuracy_score(true_label, pred_label)
    # acc = precision_score(true_label,pred_label,average='macro')
    f1_scores_median = f1_score(true_label, pred_label, average=None)
    # print(f'f1 list: {f1_scores_median}')
    f1_scores_median = np.median(f1_scores_median)
    f1_scores_macro = f1_score(true_label, pred_label, average='macro')
    f1_scores_micro = f1_score(true_label, pred_label, average='micro')
    f1_scores_weighted = f1_score(true_label, pred_label, average='weighted')
    # print('acc:', acc, 'ari:', ari, 'f1_scores_median:', f1_scores_median, 'f1_scores_macro:',
    #       f1_scores_macro, 'f1_scores_micro:', f1_scores_micro, 'f1_scores_weighted:', f1_scores_weighted)
    return acc, ari, f1_scores_median, f1_scores_macro, f1_scores_micro, f1_scores_weighted


def save_model(model, opt, model_path):
    torch.save({'model': model.state_dict(), 'opt': opt.state_dict()}, model_path)


def set_seed(seed=None):
    if seed is not None:
        os.environ['PYTHONHASHSEED'] = str(seed)
        torch.manual_seed(seed)
        torch.cuda.manual_seed(seed)  # 让显卡产生的随机数一致
        torch.cuda.manual_seed_all(seed)  # 多卡模式下，让所有显卡生成的随机数一致？这个待验证
        np.random.seed(seed)  # numpy产生的随机数一致
        random.seed(seed)
        torch.backends.cudnn.deterministic = True

    # CUDA中的一些运算，如对sparse的CUDA张量与dense的CUDA张量调用torch.bmm()，它通常使用不确定性算法。
    # 为了避免这种情况，就要将这个flag设置为True，让它使用确定的实现。
    # torch.backends.cudnn.deterministic = True

    # 设置这个flag可以让内置的cuDNN的auto-tuner自动寻找最适合当前配置的高效算法，来达到优化运行效率的问题。
    # 但是由于噪声和不同的硬件条件，即使是同一台机器，benchmark都可能会选择不同的算法。为了消除这个随机性，设置为 False
    # torch.backends.cudnn.benchmark = False


def txt_data_process(filepath, savepath='processed_result'):
    f = open(filepath)
    txt_lines = f.readlines()
    file_name, file_extension = os.path.splitext(os.path.basename(filepath))
    # import sys
    # savedStdout = sys.stdout  # 保存标准输出流
    # print_log = open(f"processed_result/{file_name}_processed.{file_extension}", "a")
    # sys.stdout = print_log
    result_dic = {}
    current_key = None
    for l in txt_lines:
        if l.startswith('test'):
            current_key = l
            # print(l)
            if current_key not in result_dic.keys():
                result_dic[current_key] = {}
        if l.startswith('best') or l.startswith('last') or l.__contains__('run time list'):
            prefix, data_list = l.split(':')
            # print(data_list)
            data_list = data_list.strip()
            # data_list = data_list.trim()
            # print(data_list)
            data_list = data_list[1:-1]
            # print(data_list)
            data_list = data_list.split(',')
            # print(filepath)
            # print(data_list)
            # print('---')
            data_list = list(map(float, data_list))
            if prefix not in result_dic[current_key].keys():
                result_dic[current_key][prefix] = data_list
            else:
                result_dic[current_key][prefix] += data_list

            # print(prefix)
            # print(f'{round(np.mean(data_list),4)}±{round(np.std(data_list),4)}')
    # print('\n')
    write_data_to_excel(result_dic, file_name, savepath)
    # for out_key, out_value in result_dic.items():
    #     print(out_key)
    #     for inner_key, inner_value in out_value.items():
    #         print(inner_key)
    #         print(f'{round(np.mean(inner_value),4)}±{round(np.std(inner_value),4)}')
    # f.close()
    # sys.stdout = savedStdout  # 恢复标准输出流
    # print_log.close()
    gc.collect()


def txt_data_process_extra(filepath, savepath='processed_result'):
    f = open(filepath)
    txt_lines = f.readlines()
    file_name, file_extension = os.path.splitext(os.path.basename(filepath))
    # import sys
    # savedStdout = sys.stdout  # 保存标准输出流
    # print_log = open(f"processed_result/{file_name}_processed.{file_extension}", "a")
    # sys.stdout = print_log
    result_dic = {}
    current_key = None
    for l in txt_lines:
        if l.startswith('train dataset') or l.startswith('test'):
            current_key = l.split('/')[-1]
            current_key, _ = os.path.splitext(os.path.basename(current_key))
            print(f'current key is: {current_key}')
            current_key = 'testfiletotal'
            # print(l)
            if current_key not in result_dic.keys():
                result_dic[current_key] = {}
        if l.startswith('best') or l.startswith('last'):
            prefix, data_list = l.split(':')
            # data_list = data_list.trim()
            data_list = data_list[1:-2]
            data_list = data_list.split(',')
            data_list = list(map(float, data_list))
            if prefix not in result_dic[current_key].keys():
                result_dic[current_key][prefix] = data_list
            else:
                result_dic[current_key][prefix] += data_list

            # print(prefix)
            # print(f'{round(np.mean(data_list),4)}±{round(np.std(data_list),4)}')
    # print('\n')
    write_data_to_excel(result_dic, file_name, savepath)
    # for out_key, out_value in result_dic.items():
    #     print(out_key)
    #     for inner_key, inner_value in out_value.items():
    #         print(inner_key)
    #         print(f'{round(np.mean(inner_value),4)}±{round(np.std(inner_value),4)}')
    # f.close()
    # sys.stdout = savedStdout  # 恢复标准输出流
    # print_log.close()
    gc.collect()


def txt_data_process_extra_scmodel(filepath, savepath='processed_result'):
    f = open(filepath)
    txt_lines = f.readlines()
    file_name, file_extension = os.path.splitext(os.path.basename(filepath))
    print(f'file name:{file_name}')
    # import sys
    # savedStdout = sys.stdout  # 保存标准输出流
    # print_log = open(f"processed_result/{file_name}_processed.{file_extension}", "a")
    # sys.stdout = print_log
    result_dic = {}
    current_key = None
    for l in txt_lines:
        if l.startswith('train dataset') or l.startswith('test'):
            if current_key:
                break
            current_key = l.split('/')[-1]
            current_key, _ = os.path.splitext(os.path.basename(current_key))
            print(f'current key is: {current_key}')
            current_key = 'testfiletotal'
            # print(l)
            if current_key not in result_dic.keys():
                result_dic[current_key] = {}
        if l.startswith('best') or l.startswith('last'):
            prefix, data_list = l.split(':')
            # data_list = data_list.trim()
            data_list = data_list[1:-2]
            data_list = data_list.split(',')
            data_list = list(map(float, data_list))
            if prefix not in result_dic[current_key].keys():
                result_dic[current_key][prefix] = data_list
            else:
                result_dic[current_key][prefix] += data_list

            # print(prefix)
            # print(f'{round(np.mean(data_list),4)}±{round(np.std(data_list),4)}')
    # print('\n')
    write_data_to_excel(result_dic, file_name, savepath)
    # for out_key, out_value in result_dic.items():
    #     print(out_key)
    #     for inner_key, inner_value in out_value.items():
    #         print(inner_key)
    #         print(f'{round(np.mean(inner_value),4)}±{round(np.std(inner_value),4)}')
    # f.close()
    # sys.stdout = savedStdout  # 恢复标准输出流
    # print_log.close()
    gc.collect()


def txt_data_process_multi_epoch(filepath, savepath='processed_result'):
    f = open(filepath)
    txt_lines = f.readlines()
    file_name, file_extension = os.path.splitext(os.path.basename(filepath))
    # import sys
    # savedStdout = sys.stdout  # 保存标准输出流
    # print_log = open(f"processed_result/{file_name}_processed.{file_extension}", "a")
    # sys.stdout = print_log
    result_dic = {}
    current_key = None
    for l in txt_lines:
        if l.startswith('test'):
            current_key = l
            # print(l)
            if current_key not in result_dic.keys():
                result_dic[current_key] = {}
        if l.startswith('best') or l.startswith('last') or l.__contains__('run time list'):
            prefix, data_list = l.split(':')
            # data_list = data_list.trim()
            data_list = data_list[1:-2]
            data_list = data_list.split(',')
            data_list = list(map(float, data_list))
            if prefix not in result_dic[current_key].keys():
                result_dic[current_key][prefix] = data_list
            else:
                result_dic[current_key][prefix] += data_list
        elif l.startswith('ctx'):
            prefix, data_list = l.split(':')
            data_list = data_list[1:-1]
            data_list = ast.literal_eval(data_list)
            result_dic[current_key][prefix] = data_list

            # print(prefix)
            # print(f'{round(np.mean(data_list),4)}±{round(np.std(data_list),4)}')
    # print('\n')
    write_data_to_excel_multi_epoch(result_dic, file_name, savepath)
    # for out_key, out_value in result_dic.items():
    #     print(out_key)
    #     for inner_key, inner_value in out_value.items():
    #         print(inner_key)
    #         print(f'{round(np.mean(inner_value),4)}±{round(np.std(inner_value),4)}')
    # f.close()
    # sys.stdout = savedStdout  # 恢复标准输出流
    # print_log.close()
    gc.collect()


def write_data_to_excel(data_dic, filename, savepath):
    w = xlwt.Workbook()
    for k, v in data_dic.items():
        worksheet = w.add_sheet(k.replace('\n', '').replace(' ', '').replace(':', ' '))
        # data_num = len(data_dic.keys())
        row_id, col_id = 0, 0
        c_row_id, c_col_id = row_id, col_id
        for name in v.keys():
            worksheet.write(c_row_id, c_col_id, f'{name} mean')
            c_col_id += 1
            worksheet.write(c_row_id, c_col_id, f'{name} std')
            c_col_id += 1
            worksheet.write(c_row_id, c_col_id, f'{name} result')
            c_col_id += 1
        row_id += 1
        for key, value in v.items():
            mean_val = round(np.mean(value), 4)
            std_val = round(np.std(value), 4)
            worksheet.write(row_id, col_id, mean_val)
            col_id += 1
            worksheet.write(row_id, col_id, std_val)
            col_id += 1
            worksheet.write(row_id, col_id, f'{mean_val}±{std_val}')
            col_id += 1
        row_id += 1
        col_id = 0
        worksheet.write(row_id, col_id, 'best acc')
        col_id += 1
        worksheet.write(row_id, col_id, 'best f1 macro')
        row_id += 1
        col_id = 0
        worksheet.write(row_id, col_id, label=xlwt.Formula('C2'))
        col_id += 1
        worksheet.write(row_id, col_id, label=xlwt.Formula('O2'))
        row_id += 1
        col_id = 0
        worksheet.write(row_id, col_id, label=xlwt.Formula('X2'))
        col_id += 1
        worksheet.write(row_id, col_id, label=xlwt.Formula('AJ2'))
        row_id += 1
        col_id = 0
        worksheet.write(row_id, col_id, 'last acc')
        col_id += 1
        worksheet.write(row_id, col_id, 'last f1 macro')

    print(f'save path:{savepath}')
    w.save(f'{savepath}/{filename}.xls')


def write_data_to_excel_multi_epoch(data_dic, filename, savepath):
    w = xlwt.Workbook()
    for k, v in data_dic.items():
        worksheet = w.add_sheet(k.replace('\n', '').replace(' ', '').replace(':', ' '))
        # data_num = len(data_dic.keys())
        row_id, col_id = 0, 0
        c_row_id, c_col_id = row_id, col_id
        for name in v.keys():
            if name.startswith('ctx'):
                continue
            worksheet.write(c_row_id, c_col_id, f'{name} mean')
            c_col_id += 1
            worksheet.write(c_row_id, c_col_id, f'{name} std')
            c_col_id += 1
            worksheet.write(c_row_id, c_col_id, f'{name} result')
            c_col_id += 1
        for name in v.keys():
            if name.startswith('ctx'):
                data_list = v[name]
                idx = 0
                worksheet.write(c_row_id, c_col_id, f'{name}_1 epoch mean')
                c_row_id += 1
                worksheet.write(c_row_id, c_col_id, f'{name}_1 epoch std')
                c_row_id += 1
                worksheet.write(c_row_id, c_col_id, f'{name}_1 epoch result')
                c_row_id += 1
                idx += 1
                while idx * 10 <= len(data_list[0]):
                    worksheet.write(c_row_id, c_col_id, f'{name}_{idx * 10} epoch mean')
                    c_row_id += 1
                    worksheet.write(c_row_id, c_col_id, f'{name}_{idx * 10} epoch std')
                    c_row_id += 1
                    worksheet.write(c_row_id, c_col_id, f'{name}_{idx * 10} epoch result')
                    c_row_id += 1
                    idx += 1

        row_id += 1
        for key, value in v.items():
            if key.startswith('ctx'):
                continue
            mean_val = round(np.mean(value), 4)
            std_val = round(np.std(value), 4)
            worksheet.write(row_id, col_id, mean_val)
            col_id += 1
            worksheet.write(row_id, col_id, std_val)
            col_id += 1
            worksheet.write(row_id, col_id, f'{mean_val}±{std_val}')
            col_id += 1
        #
        col_id += 1
        row_id = 0
        for name in v.keys():
            if name.startswith('ctx'):
                data_list = v[name]
                idx = 0
                # print(data_list[:][0])
                data_np = np.array(data_list)
                mean_val = round(np.mean(data_np[:, 0]), 4)
                std_val = round(np.std(data_np[:, 0]), 4)
                worksheet.write(row_id, col_id, mean_val)
                row_id += 1
                worksheet.write(row_id, col_id, std_val)
                row_id += 1
                worksheet.write(row_id, col_id, f'{mean_val}±{std_val}')
                row_id += 1
                idx += 1
                # print(f'data list len {len(data_list[0])}')
                while idx * 10 <= len(data_list[0]):
                    data_np = np.array(data_list)
                    # print(data_np[:, 0].shape)
                    # print(data_list[:, 0])
                    # for s in range(len(data_list)):
                    #     print(len(data_list[s]))
                    mean_val = round(np.mean(data_np[:, idx * 10 - 1]), 4)
                    std_val = round(np.std(data_np[:, idx * 10 - 1]), 4)
                    worksheet.write(row_id, col_id, mean_val)
                    row_id += 1
                    worksheet.write(row_id, col_id, std_val)
                    row_id += 1
                    worksheet.write(row_id, col_id, f'{mean_val}±{std_val}')
                    row_id += 1
                    idx += 1

        row_id = 2
        col_id = 0
        worksheet.write(row_id, col_id, 'best acc')
        col_id += 1
        worksheet.write(row_id, col_id, 'best f1 macro')
        row_id += 1
        col_id = 0
        worksheet.write(row_id, col_id, label=xlwt.Formula('C2'))
        col_id += 1
        worksheet.write(row_id, col_id, label=xlwt.Formula('O2'))
        row_id += 1
        col_id = 0
        worksheet.write(row_id, col_id, label=xlwt.Formula('X2'))
        col_id += 1
        worksheet.write(row_id, col_id, label=xlwt.Formula('AJ2'))
        row_id += 1
        col_id = 0
        worksheet.write(row_id, col_id, 'last acc')
        col_id += 1
        worksheet.write(row_id, col_id, 'last f1 macro')
    w.save(f'{savepath}/{filename}.xls')


def merge_processed_result(filepath, excel_path):
    source_wb = xlrd.open_workbook(filepath)
    target_wb = load_workbook(excel_path)

    source_filename, _ = os.path.splitext(os.path.basename(filepath))
    tissue_name = source_filename.split('printlog')[0]
    tissue_name = tissue_name[:-1]
    if 'mouse' in tissue_name:
        tissue_name = tissue_name[6:]
    if '_Extra' in tissue_name:
        tissue_name = tissue_name.split('_Extra')[0]
    elif 'Extra' in tissue_name:
        tissue_name = tissue_name.split('Extra')[0]
    print(f'tissue name:{tissue_name}')

    st_names = source_wb.sheet_names()
    for st_name in st_names:
        print(f'st name:{st_name}')
        source_sheet = source_wb.sheet_by_name(st_name)
        if st_name not in target_wb.sheetnames:
            continue
        target_sheet = target_wb[st_name]
        source_rows = source_sheet.nrows
        source_cols = source_sheet.ncols
        target_rows = target_sheet.max_row
        best_acc, best_f1_macro, last_acc, last_f1_macro = 0, 0, 0, 0
        tissue_row = 0
        for i in range(source_cols):
            # print(source_sheet.cell_value(rowx=0, colx=i))
            if source_sheet.cell_value(rowx=0, colx=i) == 'best acc list mean':
                best_acc = source_sheet.cell_value(rowx=1, colx=i)
                # print(source_sheet.cell_value(rowx=0, colx=i))
            if source_sheet.cell_value(rowx=0, colx=i) == 'best f1_scores_macro list mean':
                best_f1_macro = source_sheet.cell_value(rowx=1, colx=i)
            if source_sheet.cell_value(rowx=0, colx=i) == 'last acc list mean':
                last_acc = source_sheet.cell_value(rowx=1, colx=i)
            if source_sheet.cell_value(rowx=0, colx=i) == 'last f1_scores_macro list mean':
                last_f1_macro = source_sheet.cell_value(rowx=1, colx=i)
        for i in range(target_rows):
            print(target_sheet.cell(row=i + 1, column=1).value)
            if target_sheet.cell(row=i + 1, column=1).value == tissue_name:
                tissue_row = i + 1
        if tissue_row == 0:
            print(f'not find tissue: {tissue_name}')
            continue
        print(f'tissue row:{tissue_row}')
        acc_col, macro_col = 0, 0
        print(f'source name: {source_filename}')
        acc_col, macro_col = 5, 6
        # if 'nlp' in source_filename:
        #     print('nlp')
        #     acc_col, macro_col = 9, 10
        # elif 'nopretrain_freeze' in source_filename:
        #     print('nopretrain_freeze')
        #     acc_col, macro_col = 13, 14
        # elif 'nopretrain_nofreeze' in source_filename:
        #     print('nopretrain_nofreeze')
        #     acc_col, macro_col = 15, 16
        # elif 'pretrain_nofreeze' in source_filename:
        #     print('pretrain_nofreeze')
        #     acc_col, macro_col = 11, 12
        # elif 'pretrainnoval_finetunenoval' in source_filename:
        #     print('pretrainnoval_finetunenoval')
        #     acc_col, macro_col = 7, 8
        # elif 'pretrainval_finetunenoval' in source_filename:
        #     print('pretrainval_finetunenoval')
        #     continue
        # elif 'test_runtime' in source_filename:
        #     print('test_runtime')
        #     continue
        # else:
        #     print('no')
        #     acc_col, macro_col = 5, 6
        target_sheet.cell(row=tissue_row, column=acc_col).value = best_acc
        target_sheet.cell(row=tissue_row, column=macro_col).value = best_f1_macro
        target_sheet.cell(row=tissue_row + 1, column=acc_col).value = last_acc
        target_sheet.cell(row=tissue_row + 1, column=macro_col).value = last_f1_macro
    target_wb.save(excel_path)


def merge_processed_runtime_result_scmodel(filepath, excel_path):
    source_wb = xlrd.open_workbook(filepath)
    target_wb = load_workbook(excel_path)

    source_filename, _ = os.path.splitext(os.path.basename(filepath))
    tissue_name = source_filename.split('printlog')[0]
    tissue_name = tissue_name[:-1]
    if 'mouse' in tissue_name:
        tissue_name = tissue_name[6:]
    print(f'tissue name:{tissue_name}')

    st_names = source_wb.sheet_names()
    for st_name in st_names:
        print(f'st name:{st_name}')
        source_sheet = source_wb.sheet_by_name(st_name)
        if st_name not in target_wb.sheetnames:
            continue
        target_sheet = target_wb[st_name]
        source_rows = source_sheet.nrows
        source_cols = source_sheet.ncols
        target_rows = target_sheet.max_row
        pre_train_run_time, finetune_run_time = 0, 0
        tissue_row = 0
        for i in range(source_cols):
            print(source_sheet.cell_value(rowx=0, colx=i))
            if source_sheet.cell_value(rowx=0, colx=i) == 'pretrain run time list mean':
                pre_train_run_time = source_sheet.cell_value(rowx=1, colx=i)
                # print(source_sheet.cell_value(rowx=0, colx=i))
            if source_sheet.cell_value(rowx=0, colx=i) == 'fine tune run time list mean':
                finetune_run_time = source_sheet.cell_value(rowx=1, colx=i)
                # print(source_sheet.cell_value(rowx=0, colx=i))
        for i in range(target_rows):
            print(target_sheet.cell(row=i + 1, column=1).value)
            if target_sheet.cell(row=i + 1, column=1).value == tissue_name:
                tissue_row = i + 1
        if tissue_row == 0:
            print(f'not find tissue: {tissue_name}')
            continue
        print(f'tissue row:{tissue_row}')
        acc_col, macro_col = 0, 0
        print(f'source name: {source_filename}')
        total_runtime_col = 4
        # if 'nlp' in source_filename:
        #     print('nlp')
        #     acc_col, macro_col = 9, 10
        # elif 'nopretrain_freeze' in source_filename:
        #     print('nopretrain_freeze')
        #     acc_col, macro_col = 13, 14
        # elif 'nopretrain_nofreeze' in source_filename:
        #     print('nopretrain_nofreeze')
        #     acc_col, macro_col = 15, 16
        # elif 'pretrain_nofreeze' in source_filename:
        #     print('pretrain_nofreeze')
        #     acc_col, macro_col = 11, 12
        # elif 'pretrainnoval_finetunenoval' in source_filename:
        #     print('pretrainnoval_finetunenoval')
        #     acc_col, macro_col = 7, 8
        # elif 'pretrainval_finetunenoval' in source_filename:
        #     print('pretrainval_finetunenoval')
        #     continue
        # elif 'test_runtime' in source_filename:
        #     print('test_runtime')
        #     continue
        # else:
        #     print('no')
        #     acc_col, macro_col = 5, 6
        target_sheet.cell(row=tissue_row, column=total_runtime_col).value = pre_train_run_time + finetune_run_time
        target_sheet.cell(row=tissue_row, column=total_runtime_col + 1).value = pre_train_run_time
        target_sheet.cell(row=tissue_row, column=total_runtime_col + 2).value = finetune_run_time
    target_wb.save(excel_path)


def merge_processed_runtime_result(filepath, excel_path):
    source_wb = xlrd.open_workbook(filepath)
    target_wb = load_workbook(excel_path)

    source_filename, _ = os.path.splitext(os.path.basename(filepath))
    tissue_name = source_filename.split('printlog')[0]
    tissue_name = tissue_name[:-1]
    if 'mouse' in tissue_name:
        tissue_name = tissue_name[6:]
    print(f'tissue name:{tissue_name}')

    st_names = source_wb.sheet_names()
    for st_name in st_names:
        print(f'st name:{st_name}')
        source_sheet = source_wb.sheet_by_name(st_name)
        if st_name not in target_wb.sheetnames:
            continue
        target_sheet = target_wb[st_name]
        source_rows = source_sheet.nrows
        source_cols = source_sheet.ncols
        target_rows = target_sheet.max_row
        train_run_time = 0
        tissue_row = 0
        for i in range(source_cols):
            # print(source_sheet.cell_value(rowx=0, colx=i))
            if source_sheet.cell_value(rowx=0, colx=i) == 'run time list mean':
                train_run_time = source_sheet.cell_value(rowx=1, colx=i)
                # print(source_sheet.cell_value(rowx=0, colx=i))
        for i in range(target_rows):
            print(target_sheet.cell(row=i + 1, column=1).value)
            if target_sheet.cell(row=i + 1, column=1).value == tissue_name:
                tissue_row = i + 1
        if tissue_row == 0:
            print(f'not find tissue: {tissue_name}')
            continue
        print(f'tissue row:{tissue_row}')
        acc_col, macro_col = 0, 0
        print(f'source name: {source_filename}')
        total_runtime_col = 4
        # if 'nlp' in source_filename:
        #     print('nlp')
        #     acc_col, macro_col = 9, 10
        # elif 'nopretrain_freeze' in source_filename:
        #     print('nopretrain_freeze')
        #     acc_col, macro_col = 13, 14
        # elif 'nopretrain_nofreeze' in source_filename:
        #     print('nopretrain_nofreeze')
        #     acc_col, macro_col = 15, 16
        # elif 'pretrain_nofreeze' in source_filename:
        #     print('pretrain_nofreeze')
        #     acc_col, macro_col = 11, 12
        # elif 'pretrainnoval_finetunenoval' in source_filename:
        #     print('pretrainnoval_finetunenoval')
        #     acc_col, macro_col = 7, 8
        # elif 'pretrainval_finetunenoval' in source_filename:
        #     print('pretrainval_finetunenoval')
        #     continue
        # elif 'test_runtime' in source_filename:
        #     print('test_runtime')
        #     continue
        # else:
        #     print('no')
        #     acc_col, macro_col = 5, 6
        target_sheet.cell(row=tissue_row, column=total_runtime_col).value = train_run_time
    target_wb.save(excel_path)


def merge_processed_result_multi_epoch(filepath):
    source_wb = xlrd.open_workbook(filepath)
    target_wb = load_workbook('E:/论文/自己的论文/对比方法/实验结果模板pretrain80_nofreeze_finetune80.xlsx')

    source_filename, _ = os.path.splitext(os.path.basename(filepath))
    tissue_name = source_filename.split('printlog')[0]
    tissue_name = tissue_name[:-1]
    if 'mouse' in tissue_name:
        tissue_name = tissue_name[6:]
    print(f'tissue name:{tissue_name}')

    st_names = source_wb.sheet_names()
    for st_name in st_names:
        print(f'st name:{st_name}')
        source_sheet = source_wb.sheet_by_name(st_name)
        target_sheet = target_wb[st_name]
        source_rows = source_sheet.nrows
        source_cols = source_sheet.ncols
        target_rows = target_sheet.max_row
        best_acc, best_f1_macro, last_acc, last_f1_macro = 0, 0, 0, 0
        tissue_row = 0
        for i in range(source_cols):
            # print(source_sheet.cell_value(rowx=0, colx=i))
            if source_sheet.cell_value(rowx=0, colx=i) == 'best acc list result':
                best_acc = source_sheet.cell_value(rowx=1, colx=i)
                # print(source_sheet.cell_value(rowx=0, colx=i))
            if source_sheet.cell_value(rowx=0, colx=i) == 'best f1_scores_macro list result':
                best_f1_macro = source_sheet.cell_value(rowx=1, colx=i)
            if source_sheet.cell_value(rowx=0, colx=i) == 'last acc list result':
                last_acc = source_sheet.cell_value(rowx=1, colx=i)
            if source_sheet.cell_value(rowx=0, colx=i) == 'last f1_scores_macro list result':
                last_f1_macro = source_sheet.cell_value(rowx=1, colx=i)
        for i in range(target_rows):
            print(target_sheet.cell(row=i + 1, column=1).value)
            if target_sheet.cell(row=i + 1, column=1).value == tissue_name:
                tissue_row = i + 1
        if tissue_row == 0:
            print(f'not find tissue: {tissue_name}')
            continue
        print(f'tissue row:{tissue_row}')
        acc_col, macro_col = 0, 0
        print(f'source name: {source_filename}')
        acc_col, macro_col = 5, 6
        # if 'nlp' in source_filename:
        #     print('nlp')
        #     acc_col, macro_col = 9, 10
        # elif 'nopretrain_freeze' in source_filename:
        #     print('nopretrain_freeze')
        #     acc_col, macro_col = 13, 14
        # elif 'nopretrain_nofreeze' in source_filename:
        #     print('nopretrain_nofreeze')
        #     acc_col, macro_col = 15, 16
        # elif 'pretrain_nofreeze' in source_filename:
        #     print('pretrain_nofreeze')
        #     acc_col, macro_col = 11, 12
        # elif 'pretrainnoval_finetunenoval' in source_filename:
        #     print('pretrainnoval_finetunenoval')
        #     acc_col, macro_col = 7, 8
        # elif 'pretrainval_finetunenoval' in source_filename:
        #     print('pretrainval_finetunenoval')
        #     continue
        # elif 'test_runtime' in source_filename:
        #     print('test_runtime')
        #     continue
        # else:
        #     print('no')
        #     acc_col, macro_col = 5, 6
        target_sheet.cell(row=tissue_row, column=acc_col).value = best_acc
        target_sheet.cell(row=tissue_row, column=macro_col).value = best_f1_macro
        target_sheet.cell(row=tissue_row + 1, column=acc_col).value = last_acc
        target_sheet.cell(row=tissue_row + 1, column=macro_col).value = last_f1_macro
    target_wb.save('E:/论文/自己的论文/对比方法/实验结果模板pretrain80_nofreeze_finetune80.xlsx')


def count_key_nums(cnter_list: List[Counter], key_names):
    cnt = 0
    for key in key_names:
        for cnter in cnter_list:
            cnt += cnter.get(key, 0)
    return cnt


def draw_vnn(data_list, set_names, save_name='vnn.pdf'):
    my_dpi = 150
    plt.figure(figsize=(580 / my_dpi, 580 / my_dpi), dpi=my_dpi)
    if len(set_names) == 3:
        if isinstance(data_list[0], set):
            g = venn3(subsets=data_list, set_labels=set_names)
        else:
            data_set_list = []
            for data in data_list:
                data_set_list.append(set(data))
            g = venn3(subsets=data_set_list, set_labels=set_names)
    elif len(set_names) == 2:
        if isinstance(data_list[0], set):
            g = venn2(subsets=data_list, set_labels=set_names)
        else:
            data_set_list = []
            for data in data_list:
                data_set_list.append(set(data))
            g = venn2(subsets=data_set_list, set_labels=set_names)
    plt.show()
    plt.savefig(save_name)


def drow_vnn_with_count(data_list, set_names):
    my_dpi = 150
    plt.figure(figsize=(580 / my_dpi, 580 / my_dpi), dpi=my_dpi)
    if len(set_names) == 3:
        if isinstance(data_list[0], set):
            g = venn3(subsets=data_list, set_labels=set_names)
        else:
            data_set_list = []
            data_counter_list = []
            for data in data_list:
                data_set_list.append(set(data))
                data_counter_list.append(Counter(data))
            key_111 = data_set_list[0] & data_set_list[1] & data_set_list[2]
            key_110_111 = data_set_list[0] & data_set_list[1]
            key_101_111 = data_set_list[0] & data_set_list[2]
            key_011_111 = data_set_list[1] & data_set_list[2]
            key_101 = key_101_111 - key_111
            key_110 = key_110_111 - key_111
            key_011 = key_011_111 - key_111
            key_100 = data_set_list[0] - key_101 - key_110_111
            key_010 = data_set_list[1] - key_011 - key_110_111
            key_001 = data_set_list[2] - key_101 - key_011_111

            cnt_100 = count_key_nums([data_counter_list[0]], key_100)
            cnt_101 = count_key_nums([data_counter_list[0], data_counter_list[2]], key_101)
            cnt_110 = count_key_nums([data_counter_list[0], data_counter_list[1]], key_110)
            cnt_111 = count_key_nums([data_counter_list[0], data_counter_list[1], data_counter_list[2]], key_111)
            cnt_001 = count_key_nums([data_counter_list[2]], key_001)
            cnt_010 = count_key_nums([data_counter_list[1]], key_010)
            cnt_011 = count_key_nums([data_counter_list[1], data_counter_list[2]], key_011)

            g = venn3(subsets=data_set_list, set_labels=set_names)
            g.get_label_by_id('100').set_text(str(cnt_100))
            g.get_label_by_id('101').set_text(str(cnt_101))
            g.get_label_by_id('110').set_text(str(cnt_110))
            g.get_label_by_id('111').set_text(str(cnt_111))
            g.get_label_by_id('001').set_text(str(cnt_001))
            g.get_label_by_id('010').set_text(str(cnt_010))
            g.get_label_by_id('011').set_text(str(cnt_011))

    plt.show()
    plt.savefig('vnn_counter.pdf')


def draw_heat_map(X, x_label, y_label, title, save_path):
    my_dpi = 300
    # plt.figure(figsize=(580 / my_dpi, 580 / my_dpi), dpi=my_dpi)
    plt.figure(figsize=(10,10), dpi=my_dpi)
    sns.heatmap(X, cmap='OrRd', xticklabels=x_label, yticklabels=y_label)
    # plt.xticks(fontsize=1)
    # plt.yticks(fontsize=4)

    # my_dpi = 150
    # fig=plt.figure(figsize=(580 / my_dpi, 580 / my_dpi), dpi=my_dpi)
    # ax = fig.add_subplot(111)
    #
    # ax.set_yticks(range(len(y_label)))
    # ax.set_yticklabels(y_label)
    # ax.set_xticks(range(len(x_label)))
    # ax.set_xticklabels(x_label)
    #
    # im = ax.imshow(X, cmap=plt.cm.hot_r)
    # plt.colorbar(im)

    plt.title(title)
    # _, ax = plt.subplots()
    # ax.tick_params(axis='x', labelsize=8)
    # ax.tick_params(axis='y', labelsize=8)
    # plt.tight_layout()
    # plt.gcf().subplots_adjust(left=0.05, top=0.9, bottom=0.1)

    plt.show()
    plt.savefig(save_path, bbox_inches='tight')
    plt.close()


def draw_pie_scatter(embedding, prop, save_file_name, cell_type_idx_dic, color_map):
    reducer = manifold.TSNE(n_components=2, init='pca', random_state=1)
    embedding = reducer.fit_transform(embedding)

    cell_num = prop.shape[0]
    type_num = prop.shape[1]
    new_ss = np.zeros_like(prop)
    prop[prop < 0.05] = 0
    prop = prop / prop.sum(-1, keepdims=True)
    for i in range(type_num):
        new_ss[:, i] = prop[:, 0:i + 1].sum(-1)
    print(f'cell num:{cell_num}')
    print(new_ss)
    plt.figure(figsize=(10, 10), dpi=300)
    for i in range(cell_num):
        print(f'cur cell:{i + 1}')
        pre, cur = 0, 0
        # xy_total = []
        for j in range(type_num):
            pre = cur
            cur = 2 * np.pi * new_ss[i][j]
            if pre == cur:
                continue
            x = [0] + np.cos(np.linspace(pre, cur, 15)).tolist()
            y = [0] + np.sin(np.linspace(pre, cur, 15)).tolist()
            xy = list(zip(x, y))
            # xy_total.append(xy)
            # for j in range(type_num):
            label_name = cell_type_idx_dic.getGene(j)
            color, _ = color_map[label_name]
            plt.scatter(embedding[i, 0], embedding[i, 1], marker=(xy), s=100, c=color)
            # plt.scatter(embedding[i, 0], embedding[i, 1], marker=(xy))
        # if i+1==100:
        #     break
    plt.show()
    plt.savefig(save_file_name)
    plt.close()
    plt.clf()


def count_cell_num_dic(filepath):
    adata = check_anndata(filepath)
    print(adata.uns_keys())
    cell_num_Counter = Counter(adata.obs['cell_type'])
    print(cell_num_Counter)


def check_scmodel_extra():
    file_list = glob.glob(
        # 'impl/extra_result/*_pretrain40epoch_finetune40epoch_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout_new.txt')
        'impl/Ablation/*_printlog_pretrain40epoch_finetune40epoch_2layer_2head_meanval_dot_embedding_nomaskenhance_noembeddingdropout.txt')

    print(file_list)
    for f in file_list:
        # txt_data_process(f, savepath='impl/final_result/lab/processed_result')
        txt_data_process(f, savepath='impl/Ablation/process')
        # txt_data_process_extra_scmodel(f, savepath='impl/final_result/lab/processed_result')
        # txt_data_process_extra_scmodel(f, savepath='impl/extra_result/processed_result')

    file_list = glob.glob(
        'impl/Ablation/process\*_printlog_pretrain40epoch_finetune40epoch_2layer_2head_meanval_dot_embedding_nomaskenhance_noembeddingdropout.xls')
    # file_list = glob.glob(
    #     'D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\PBMC*')
    print(file_list)
    for f in file_list:
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scmodel\实验结果模板pretrain40_finetune40_nopcainitial_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout_final.xlsx")
        merge_processed_result(f,
                               "E:\论文\自己的论文\对比方法\scmodel\实验结果模板pretrain40_finetune40_2layer_2head_val_dot_embedding_nomaskenhance_noembeddingdropout_final.xlsx")

        # if 'Extra' in f:
        #     merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scmodel\extra\Mouse_Pancreas_Task2\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout_new.xlsx")
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scmodel\extra\Mouse_Pancreas_Task2\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embedding_maskenhance08_embeddingdropout02.xlsx")
        # "E:\论文\自己的论文\对比方法\scmodel\extra\PBMC45k\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embeddding_maskenhance08_noembeddingdropout.xlsx")
        # "E:\论文\自己的论文\对比方法\scmodel\extra\PBMC45k_Task2\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embeddding_nomaskenhance_noembeddingdropout.xlsx")
        # "E:\论文\自己的论文\对比方法\scmodel\extra\PBMC45k\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embeddding_nomaskenhance_embeddingdropout01.xlsx")
        # "E:\论文\自己的论文\对比方法\scmodel\extra\Mouse_Pancreas_Task2\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout.xlsx")
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scmodel\extra\Mouse_Brain_Task2\Mouse_Brain实验结果结果pretrain40epoch_finetune40epoch_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout.xlsx")
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scmodel\extra\Mouse_Brain_Task2\Mouse_Brain实验结果结果pretrain40epoch_finetune40epoch_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout.xlsx")


def check_TOSICA_extra():
    file_list = glob.glob(
        # 'E:\pyproject\TOSICA-main\extra_result\*.txt')
        'E:\pyproject\TOSICA-main\extra_result\*.txt')

    print(file_list)
    for f in file_list:
        txt_data_process_extra(f, savepath='E:\pyproject\TOSICA-main\extra_result\processed_result')
        # txt_data_process_extra_scmodel(f, savepath='impl/extra_result/processed_result')

    file_list = glob.glob('E:\pyproject\TOSICA-main\extra_result\processed_result\*.xls')
    # file_list = glob.glob(
    #     'D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\PBMC*')
    print(file_list)
    for f in file_list:
        # if 'Extra' in f:
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\TOSICA\实验结果extra_Brain_task2.xlsx")
        merge_processed_result(f,
                               "E:\论文\自己的论文\对比方法\TOSICA\实验结果extra_Pancreas.xlsx")
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\TOSICA\实验结果extra_PBMC45k_task2.xlsx")
        # if 'Extra' in f:
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scmodel\extra\PBMC45k\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embeddding_nomaskenhance_noembeddingdropout.xlsx")


def check_Concerto_extra():
    file_list = glob.glob(
        # 'E:\pyproject\TOSICA-main\extra_result\*.txt')
        'E:\论文\contrastive learning on multi single-cell atlas\Concerto-reproducibility-main\Concerto-reproducibility-main\extra_result\*.txt')

    print(file_list)
    for f in file_list:
        txt_data_process_extra(f,
                               savepath='E:\论文\contrastive learning on multi single-cell atlas\Concerto-reproducibility-main\Concerto-reproducibility-main\extra_result\processed_result')
        # txt_data_process_extra_scmodel(f, savepath='impl/extra_result/processed_result')

    file_list = glob.glob(
        'E:\论文\contrastive learning on multi single-cell atlas\Concerto-reproducibility-main\Concerto-reproducibility-main\extra_result\processed_result\*.xls')
    # file_list = glob.glob(
    #     'D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\PBMC*')
    print(file_list)
    for f in file_list:
        # if 'Extra' in f:
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\Concerto\实验结果extra_Brain_task2.xlsx")
        merge_processed_result(f,
                               "E:\论文\自己的论文\对比方法\Concerto\实验结果extra_Pancreas.xlsx")
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\Concerto\实验结果extra_PBMC45k_task2.xlsx")


def check_ScSemiGAN_extra():
    file_list = glob.glob(
        'E:\论文\scSemiGan\scSemiGAN-main\scSemiGAN-main\large\*.txt')

    print(file_list)
    for f in file_list:
        txt_data_process_extra(f, savepath='E:\论文\scSemiGan\scSemiGAN-main\scSemiGAN-main\large\processed_result')
        # txt_data_process_extra_scmodel(f, savepath='impl/extra_result/processed_result')

    # file_list = glob.glob('E:\论文\scSemiGan\scSemiGAN-main\scSemiGAN-main\extra_result\processed_result\*.xls')
    # file_list = glob.glob(
    #     'D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\PBMC*')
    # print(file_list)
    # for f in file_list:
    # if 'Extra' in f:
    # merge_processed_result(f,
    #                        "E:\论文\自己的论文\对比方法\scSemiGAN\实验结果extra_Brain_Task2.xlsx")
    # merge_processed_result(f,
    #                        "E:\论文\自己的论文\对比方法\scSemiGAN\实验结果extra_Pancreas.xlsx")
    # merge_processed_result(f,
    #                        "E:\论文\自己的论文\对比方法\scSemiGAN\实验结果extra_PBMC45k_Task2.xlsx")


def check_itclust_extra():
    file_list = glob.glob(
        'E:\论文\itclust\ItClust-master\\tutorial\large\*.txt')

    print(file_list)
    for f in file_list:
        txt_data_process_extra(f, savepath='E:\论文\itclust\ItClust-master\\tutorial\large\processed_result')
        # txt_data_process_extra_scmodel(f, savepath='impl/extra_result/processed_result')

    # file_list = glob.glob('E:\论文\itclust\ItClust-master\\tutorial\extra_result\processed_result\*.xls')
    # file_list = glob.glob(
    #     'D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\PBMC*')
    # print(file_list)
    # for f in file_list:
    # if 'Extra' in f:
    # merge_processed_result(f,
    #                        "E:\论文\自己的论文\对比方法\itClust\实验结果extra_Brain_task2.xlsx")
    # merge_processed_result(f,
    #                        "E:\论文\自己的论文\对比方法\itClust\实验结果extra_Pancreas.xlsx")
    # merge_processed_result(f,
    #                        "E:\论文\自己的论文\对比方法\itClust\实验结果extra_PBMC45k_task2.xlsx")


def check_scdeepsort_extra():
    file_list = glob.glob(
        'D:\代码\scDeepSort-master\\result\extra\\task2\*.txt')

    print(file_list)
    for f in file_list:
        txt_data_process_extra(f, savepath='D:\代码\scDeepSort-master\\result\extra\\task2\processed_result')
        # txt_data_process_extra_scmodel(f, savepath='impl/extra_result/processed_result')

    file_list = glob.glob('D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\*.xls')
    # file_list = glob.glob(
    #     'D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\PBMC*')
    print(file_list)
    for f in file_list:
        # if 'Extra' in f:
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scDeepSort\实验结果extra_Brain_task2.xlsx")
        merge_processed_result(f,
                               "E:\论文\自己的论文\对比方法\scDeepSort\实验结果extra_Pancreas_task2.xlsx")
        # merge_processed_result(f,
        #                        "E:\论文\自己的论文\对比方法\scDeepSort\实验结果extra_PBMC45k_task2.xlsx")


if __name__ == '__main__':
    check_scmodel_extra()
    # check_scdeepsort_extra()
    # check_Concerto_extra()
    # check_TOSICA_extra()
    # check_ScSemiGAN_extra()
    # check_itclust_extra()
    # tsne_plot()
    # count_cell_num_dic('../../datasets/mouse/Spleen/mouse_Spleen.h5ad')

    # txt_data_process_multi_epoch('C:\\Users\zzydh\Desktop\\test.txt')
    # file_list = glob.glob('impl\labresult\pretrain80_fintune80_multiepoch\*.txt')
    # for f in file_list:
    #     txt_data_process_multi_epoch(f, savepath='impl\labresult\pretrain80_fintune80_multiepoch\processed_result')

    # file_list = glob.glob('E:\pyproject\scmodel\models\gat\impl\labresult\pretrain80_fintune80_multiepoch\*.txt')
    # for f in file_list:
    #     txt_data_process(f, savepath='E:\pyproject\scmodel\models\gat\impl\labresult\pretrain80_fintune80_multiepoch\processed_result')
    #
    #
    # # file_list = glob.glob('E:\pyproject\scmodel\models\gat\impl\labresult\pretrain80_fintune80_multiepoch\processed_result\*testsize095*.xls')
    # # for f in file_list:
    # #     merge_processed_result(f)

    # # file_list = glob.glob('E:\论文\contrastive learning on multi single-cell atlas\Concerto-reproducibility-main\Concerto-reproducibility-main\\finalresult\*.txt')
    # # for f in file_list:
    # #     txt_data_process(f, savepath='E:\论文\contrastive learning on multi single-cell atlas\Concerto-reproducibility-main\Concerto-reproducibility-main\processed_result')
    #
    #
    # file_list = glob.glob('E:\pyproject\scmodel\models\gat\impl\\final_result\lab\*1layer_1head_meanval_dot_embedding_nomaskenhance*.xls')
    # for f in file_list:
    #     merge_processed_result(f, 'E:\论文\自己的论文\对比方法\实验结果Concerto.xlsx')

    # file_list = glob.glob('D:\代码\scDeepSort-master\\result\extra\\task2\*.txt')
    # for f in file_list:
    #     txt_data_process_extra(f, savepath='D:\代码\scDeepSort-master\\result\extra\\task2\processed_result')
    #

    # file_list = glob.glob('impl\\final_result\lab\*printlog_pretrain40epoch_finetune40epoch_1layer_1head_val_dot_embedding_nomaskenhance_noembeddingdropout_final*')
    # file_list = glob.glob(
    #     'impl\\extra_result\*_printlog_pretrain40epoch_finetune40epoch_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout_fix.txt')
    #
    # print(file_list)
    # for f in file_list:
    #     # txt_data_process(f, savepath='impl/final_result\lab\processed_result')
    #     txt_data_process_extra_scmodel(f, savepath='impl/extra_result/processed_result')
    #
    # file_list = glob.glob('D:\代码\scDeepSort-master\\result\extra\\task1\*.txt')
    # print(file_list)
    # for f in file_list:
    #     txt_data_process_extra(f, savepath='D:\代码\scDeepSort-master\\result\extra\\task1\processed_result')
    # file_list = glob.glob(
    #     'impl/final_result\lab\processed_result\*printlog_pretrain40epoch_finetune40epoch_1layer_1head_val_dot_embedding_nomaskenhance_embeddingdropout02_final.xls')
    #
    # file_list = glob.glob('impl/extra_result\processed_result\*_printlog_pretrain40epoch_finetune40epoch_1layer_1head_meanval_dot_embedding_nomaskenhance_noembeddingdropout_fix.xls')
    # # file_list = glob.glob(
    # #     'D:\代码\scDeepSort-master\\result\extra\\task2\processed_result\PBMC*')
    # print(file_list)
    # for f in file_list:
    #     # if 'Extra' in f:
    #     # merge_processed_result(f,
    #     #                        "E:\论文\自己的论文\对比方法\scDeepSort\实验结果extra_PBMC45k_task2.xlsx")
    #     # if 'Extra' in f:
    #     merge_processed_result(f,
    #                            "E:\论文\自己的论文\对比方法\scmodel\extra\PBMC45k\实验结果结果pretrain40epoch_finetune40epoch_on_trainset_1layer_1head_meanval_dot_embeddding_nomaskenhance_noembeddingdropout.xlsx")

    # file_list = glob.glob('impl\\final_result\RTX2080\*.txt')
    # for f in file_list:
    #     txt_data_process(f, savepath='impl\\final_result\RTX2080\\processed_result')

    # file_list = glob.glob('E:\论文\itclust\ItClust-master\\tutorial\extra_result\*.txt')
    # for f in file_list:
    #     txt_data_process_extra(f, savepath='E:\论文\itclust\ItClust-master\\tutorial\extra_result\processed_result')

    # file_list = glob.glob('E:\论文\itclust\ItClust-master\\tutorial\extra_result\processed_result\*Baron*.xls')
    # print(file_list)
    # for f in file_list:
    #     merge_processed_result(f,
    #                            "E:\论文\自己的论文\对比方法\itClust\实验结果extra_Pancreas.xlsx")
    #
    # file_list = glob.glob('impl/final_result/RTX2080/processed_result\*mean*.xls')
    # print(file_list)
    # for f in file_list:
    #     merge_processed_runtime_result_scmodel(f,
    #                            "E:\论文\自己的论文\对比方法\scmodel\\runtime\实验运行时间scmodel.xlsx")

    # file_list = glob.glob('D:\代码\scDeepSort-master\\processed_result\*.xls')
    # print(file_list)
    # for f in file_list:
    #     merge_processed_runtime_result(f,
    #                            "E:\论文\自己的论文\对比方法\scDeepSort\实验运行时间scDeepSort.xlsx")

    # data = '[[0.7674418604651163, 0.7995188452285485, 0.8083400160384924, 0.8131515637530072, 0.8159582999198075, 0.8111467522052928, 0.8167602245388933, 0.8039294306335204, 0.8175621491579792, 0.8227746591820368, 0.8095429029671211, 0.8311948676824379, 0.8195669607056937, 0.826383319967923, 0.8380112269446672, 0.8376102646351243, 0.8392141138732959, 0.8299919807538091, 0.8219727345629511, 0.8416198877305533, 0.8251804330392943, 0.834803528468324, 0.8311948676824379, 0.8360064153969526, 0.8372093023255814, 0.8279871692060946, 0.8384121892542101, 0.8456295108259824, 0.8227746591820368, 0.8480352846832397, 0.8360064153969526, 0.8323977546110666, 0.8504410585404972, 0.826383319967923, 0.8191659983961508, 0.8536487570168404, 0.8340016038492382, 0.838813151563753, 0.8227746591820368, 0.8536487570168404, 0.8504410585404972, 0.8420208500400962, 0.8368083400160385, 0.851242983159583, 0.8628708901363272, 0.8536487570168404, 0.8448275862068966, 0.8303929430633521, 0.8588612670408982, 0.8520449077786688, 0.859262229350441, 0.8352044907778668, 0.8564554931836408, 0.8319967923015237, 0.8680834001603849, 0.8668805132317562, 0.8396150761828388, 0.863672814755413, 0.8520449077786688, 0.8580593424218124, 0.8540497193263833, 0.8412189254210104, 0.8504410585404972, 0.8524458700882117, 0.8600641539695268, 0.8508420208500401, 0.8556535685645549, 0.8255813953488372, 0.8464314354450682, 0.8528468323977546, 0.8644747393744988, 0.8516439454691259, 0.8520449077786688, 0.8556535685645549, 0.8516439454691259, 0.8424218123496391, 0.8436246992782679, 0.8496391339214114, 0.8584603047313553, 0.8628708901363272], [0.7846832397754611, 0.7979149959903769, 0.8095429029671211, 0.7999198075380914, 0.8055332798716921, 0.809943865276664, 0.8127506014434643, 0.8079390537289495, 0.8147554129911788, 0.8107457898957497, 0.8163592622293504, 0.8239775461106656, 0.8207698476343224, 0.8295910184442662, 0.8207698476343224, 0.8340016038492382, 0.8151563753007217, 0.8404170008019246, 0.8404170008019246, 0.842822774659182, 0.8267842822774659, 0.8291900561347233, 0.8171611868484362, 0.8444266238973537, 0.8580593424218124, 0.8219727345629511, 0.8556535685645549, 0.851242983159583, 0.8392141138732959, 0.851242983159583, 0.8336006415396953, 0.8632718524458701, 0.8580593424218124, 0.8516439454691259, 0.855252606255012, 0.8456295108259824, 0.8628708901363272, 0.8580593424218124, 0.8600641539695268, 0.8612670408981555, 0.855252606255012, 0.8672814755412991, 0.8616680032076984, 0.8532477947072975, 0.8744987971130713, 0.8660785886126704, 0.8656776263031275, 0.8576583801122695, 0.8608660785886126, 0.8632718524458701, 0.843223736968725, 0.8508420208500401, 0.8668805132317562, 0.8628708901363272, 0.8564554931836408, 0.8728949478748997, 0.8664795509222133, 0.8680834001603849, 0.872093023255814, 0.872093023255814, 0.8556535685645549, 0.843223736968725, 0.8460304731355253, 0.8660785886126704, 0.8299919807538091, 0.8744987971130713, 0.8608660785886126, 0.8708901363271853, 0.8584603047313553, 0.8456295108259824, 0.8696872493985566, 0.8688853247794708, 0.8656776263031275, 0.8604651162790697, 0.8664795509222133, 0.8484362469927826, 0.8600641539695268, 0.867682437850842, 0.8740978348035284, 0.867682437850842], [0.7297514033680834, 0.8155573376102646, 0.8095429029671211, 0.8111467522052928, 0.8067361668003208, 0.8095429029671211, 0.8103448275862069, 0.7894947874899759, 0.8087409783480353, 0.809943865276664, 0.8055332798716921, 0.8203688853247795, 0.8131515637530072, 0.8315958299919808, 0.8239775461106656, 0.8199679230152366, 0.8400160384923817, 0.8368083400160385, 0.8055332798716921, 0.8255813953488372, 0.8247794707297514, 0.7802726543704892, 0.8323977546110666, 0.8356054530874097, 0.8011226944667201, 0.7991178829190057, 0.8275862068965517, 0.8239775461106656, 0.8372093023255814, 0.8440256615878108, 0.8352044907778668, 0.7762630312750601, 0.8319967923015237, 0.8424218123496391, 0.8384121892542101, 0.8452285485164395, 0.8340016038492382, 0.818364073777065, 0.7959101844426624, 0.8504410585404972, 0.8368083400160385, 0.8352044907778668, 0.8392141138732959, 0.8344025661587811, 0.8508420208500401, 0.8071371291098637, 0.8283881315156375, 0.8576583801122695, 0.8195669607056937, 0.8299919807538091, 0.8380112269446672, 0.8508420208500401, 0.8384121892542101, 0.8584603047313553, 0.8500400962309543, 0.8480352846832397, 0.859663191659984, 0.8532477947072975, 0.8303929430633521, 0.8323977546110666, 0.8688853247794708, 0.8648757016840417, 0.8620689655172413, 0.8640737770649559, 0.8528468323977546, 0.8492381716118684, 0.8444266238973537, 0.8524458700882117, 0.8436246992782679, 0.801523656776263, 0.8496391339214114, 0.8560545308740979, 0.830793905372895, 0.8520449077786688, 0.8380112269446672, 0.8155573376102646, 0.8119486768243785, 0.8532477947072975, 0.8291900561347233, 0.8604651162790697], [0.7594226142742582, 0.7602245388933441, 0.8035284683239775, 0.8083400160384924, 0.8119486768243785, 0.8187650360866079, 0.8311948676824379, 0.8179631114675221, 0.834803528468324, 0.8380112269446672, 0.818364073777065, 0.8380112269446672, 0.8344025661587811, 0.8331996792301524, 0.8311948676824379, 0.8211708099438653, 0.838813151563753, 0.8147554129911788, 0.8219727345629511, 0.8179631114675221, 0.8239775461106656, 0.846832397754611, 0.8464314354450682, 0.8504410585404972, 0.8311948676824379, 0.8239775461106656, 0.8492381716118684, 0.8303929430633521, 0.8488372093023255, 0.8167602245388933, 0.8087409783480353, 0.8536487570168404, 0.8572574178027266, 0.8672814755412991, 0.8604651162790697, 0.8456295108259824, 0.8291900561347233, 0.843223736968725, 0.863672814755413, 0.8436246992782679, 0.8648757016840417, 0.846832397754611, 0.8448275862068966, 0.8620689655172413, 0.8532477947072975, 0.846832397754611, 0.8564554931836408, 0.8612670408981555, 0.867682437850842, 0.8488372093023255, 0.8600641539695268, 0.8664795509222133, 0.8580593424218124, 0.8744987971130713, 0.8608660785886126, 0.859262229350441, 0.8400160384923817, 0.8664795509222133, 0.8688853247794708, 0.8700882117080995, 0.8315958299919808, 0.8704891740176424, 0.8644747393744988, 0.8600641539695268, 0.8600641539695268, 0.8548516439454691, 0.8748997594226142, 0.8448275862068966, 0.855252606255012, 0.8187650360866079, 0.8644747393744988, 0.8648757016840417, 0.8548516439454691, 0.8632718524458701, 0.8392141138732959, 0.8692862870890137, 0.8640737770649559, 0.8588612670408982, 0.8588612670408982, 0.8584603047313553], [0.7762630312750601, 0.7546110665597434, 0.7842822774659182, 0.7939053728949479, 0.7902967121090617, 0.7846832397754611, 0.7987169206094628, 0.8203688853247795, 0.8095429029671211, 0.8227746591820368, 0.8295910184442662, 0.8259823576583801, 0.8275862068965517, 0.8396150761828388, 0.8299919807538091, 0.8259823576583801, 0.8115477145148356, 0.8311948676824379, 0.8344025661587811, 0.8368083400160385, 0.8364073777064955, 0.8344025661587811, 0.8191659983961508, 0.8396150761828388, 0.8384121892542101, 0.8231756214915797, 0.8279871692060946, 0.8376102646351243, 0.8480352846832397, 0.8344025661587811, 0.8440256615878108, 0.8195669607056937, 0.842822774659182, 0.8416198877305533, 0.8083400160384924, 0.8492381716118684, 0.847233360064154, 0.838813151563753, 0.859663191659984, 0.8564554931836408, 0.8460304731355253, 0.8456295108259824, 0.8540497193263833, 0.8440256615878108, 0.8536487570168404, 0.8368083400160385, 0.8195669607056937, 0.8628708901363272, 0.8584603047313553, 0.8600641539695268, 0.863672814755413, 0.8580593424218124, 0.8640737770649559, 0.8584603047313553, 0.8440256615878108, 0.8548516439454691, 0.8532477947072975, 0.8520449077786688, 0.8560545308740979, 0.859663191659984, 0.8652766639935846, 0.8508420208500401, 0.8648757016840417, 0.847233360064154, 0.8584603047313553, 0.8648757016840417, 0.8632718524458701, 0.8628708901363272, 0.8684843624699278, 0.8311948676824379, 0.822373696872494, 0.8688853247794708, 0.871692060946271, 0.8484362469927826, 0.8604651162790697, 0.859663191659984, 0.8600641539695268, 0.8604651162790697, 0.8632718524458701, 0.8572574178027266]]'
    # data_json = ast.literal_eval(data)
    # print(data_json[0])

    # wb = xlrd.open_workbook('E:/论文/自己的论文/对比方法/消融实验.xlsx')
    # table = wb.sheet_by_index(0)
    # row = table.nrows
    # print(row)
    # for i in range(row):
    #     print(table.row_values(i))
    # from xlutils.copy import copy
    # c_wb = copy(wb)
    # wb = load_workbook('processed_result/Muscle_printlog.xls')
    # # wb = load_workbook('E:/论文/自己的论文/对比方法/消融实验.xlsx')
    # ws = wb['testsize 0.9']
    # print(ws['A1'].value)
    # print(ws['A2'].value)
    # print(ws['A3'].value)
    # print(ws['A4'].value)
    # dir_name = 'predict'
    # dataset_name = ''
    # adata_filename = ''
    # adata = check_anndata('../../datasets/mouse/Muscle/mouse_Muscle_total.h5ad')
    # pca = PCA(n_components=50)
    # pca_data = pca.fit_transform(adata.X)
    # # pca_data = adata.X
    # tsne_plot(pca_data, adata.obs['cell_type'], 'Tse_Testis.jpg')
